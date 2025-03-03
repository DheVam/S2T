from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
from openpyxl import load_workbook
import nltk
from nltk.stem import WordNetLemmatizer
from nltk.corpus import wordnet
import re
import os
import logging
import traceback
from openpyxl.styles import Border, Side, Alignment, Font

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# NLTK Setup
nltk.download('punkt')
nltk.download('wordnet')
nltk.download('averaged_perceptron_tagger')

app = Flask(__name__)
CORS(app)  # Enable CORS for frontend integration

# Define upload and output folder paths
UPLOAD_FOLDER = "uploads"
OUTPUT_FOLDER = "output"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Initialize lemmatizer
lemmatizer = WordNetLemmatizer()
stop_words = {"and", "for", "of", "by", "be", "the", "with", "in", "on", "at", "a", "an", "is", "to", "from", "or"}

def get_wordnet_pos(tag):
    """Convert NLTK POS tags to WordNet format."""
    if tag.startswith('J'):
        return wordnet.ADJ  # Adjective
    elif tag.startswith('V'):
        return wordnet.VERB  # Verb (Fix for 'purchasing' issue)
    elif tag.startswith('N'):
        return wordnet.NOUN  # Noun
    elif tag.startswith('R'):
        return wordnet.ADV  # Adverb
    else:
        return wordnet.NOUN  # Default case


def lemmatize_text_field(text):
    """Lemmatizes words in a given text field by trying multiple POS tags."""
    words = re.findall(r'\b\w+\b', text.lower())  # Extract words
    lemmatized_words = []
    
    for word in words:
        lemma = lemmatizer.lemmatize(word, wordnet.VERB)  # First try as a verb
        if lemma == word:  # If no change, try as a noun
            lemma = lemmatizer.lemmatize(word, wordnet.NOUN)
        if lemma == word:  # If still no change, try as an adjective
            lemma = lemmatizer.lemmatize(word, wordnet.ADJ)
        
        lemmatized_words.append(lemma)

    return lemmatized_words


def reconstruct_and_remove_stop_words(description, naming_dict):
    """Processes 'Business Description' to standardize and rename fields based on Naming Standards."""
    if not isinstance(description, str) or not description.strip():
        return ""

    words = lemmatize_text_field(description)  # Step 1: Lemmatization
    words = [word for word in words if word not in stop_words]  # Step 2: Remove stop words

    # Step 3: Reorder based on specific keywords ('for', 'of' should be moved to end)
    special_keywords = {"for", "of"}
    ordered_words = [word for word in words if word not in special_keywords] + \
                    [word for word in words if word in special_keywords]

    # Step 4: Map words using naming_dict
    mapped_words = [naming_dict.get(word, word) for word in ordered_words]

    # Step 5: Detect GUID/UID and enforce '_id' suffix
    if any(word in {"guid", "uid", "id"} for word in mapped_words):
        return "_".join(mapped_words) + "_id"

    return "_".join(mapped_words)

def fix_t_ts_columns(df):
    """Fixes incorrect target column names by replacing '_t_ts' with '_ts'."""
    if 'Target Column' in df.columns:
        df['Target Column'] = df['Target Column'].str.replace(r'_t_ts\b', '_ts', regex=True)
    return df


def generate_target_table_name(source_name, naming_dict):
    """Generates the target table name based on the naming dictionary."""
    if pd.isna(source_name) or not isinstance(source_name, str):
        return ""

    parts = source_name.split("_")[1:]
    target_parts = [naming_dict.get(part.lower(), part.lower()) for part in parts]

    return "_".join(target_parts)

def apply_excel_formatting(file_path):
    """Applies formatting to the Excel output file."""
    wb = load_workbook(file_path)
    ws = wb.active

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Apply formatting to header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Apply formatting to all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

    # Auto-adjust column widths based on content
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2  # Add some padding

    # Save the formatted workbook
    wb.save(file_path)
    logging.info("Excel formatting applied successfully!")

def process_files(s2t_path, naming_standards_path, output_path):
    """Processes input files and generates an updated S2T file with target column and table names, ensuring proper formatting."""
    try:
        logging.info("Reading input files...")
        s2t_df = pd.read_excel(s2t_path, engine='openpyxl')
        naming_standards_df = pd.read_excel(naming_standards_path, engine='openpyxl')

        if 'Real Meaning' not in naming_standards_df.columns:
            raise ValueError("Naming Standards file must contain a 'Real Meaning' column.")

        logging.info("Creating naming dictionary...")
        naming_dict = {
            str(row['Real Meaning']).strip().lower(): str(row.iloc[1]).strip()
            for _, row in naming_standards_df.iterrows() if len(row) >= 2
        }

        logging.info("Processing Target Columns...")
        if 'Business Description' in s2t_df.columns:
            s2t_df['Target Column'] = s2t_df['Business Description'].apply(
                lambda x: reconstruct_and_remove_stop_words(x, naming_dict)
            )

        logging.info("Processing Target Table Names...")
        if 'Source Table Name' in s2t_df.columns:
            s2t_df['Target Table'] = s2t_df['Source Table Name'].apply(
                lambda x: generate_target_table_name(x, naming_dict)
            )

        # Ensure output format matches the expected structure
        expected_columns = [
            "S.No", "Target Schema", "Review Target Schema", "Target Table", "Review Target Table",
            "Target Column", "Review Target Column", "GDA COMMENTS", "GDA REMOVE",
            "Datatype", "Review Datatype", "Constraint", "Review Constraint",
            "Mapping", "Review Mapping", "Source Table Name", "Review Source Table Name",
            "Sample Data in table", "Source Column Name", "Review Source Column Name",
            "Business Description", "Comments"
        ]

        # Add missing columns with empty values
        for col in expected_columns:
            if col not in s2t_df.columns:
                s2t_df[col] = ""

        # Reorder columns to match expected format
        s2t_df = s2t_df[expected_columns]

        logging.info("Saving processed file...")
        s2t_df.to_excel(output_path, index=False, engine='openpyxl')

        # Apply formatting
        apply_excel_formatting(output_path)

        return output_path
    except Exception as e:
        logging.error("Error processing files: %s", traceback.format_exc())
        raise Exception(f"Error processing files: {str(e)}")



@app.route('/')
def home():
    """Home route to check if the server is running."""
    return "Flask Backend is Running!"

@app.route('/process', methods=['POST'])
def upload_files():
    """Handles file upload and initiates processing."""
    logging.info("Received file upload request.")

    if 's2t_file' not in request.files or 'naming_standards_file' not in request.files:
        logging.error("Missing required files in the request.")
        return jsonify({"error": "Both S2T file and Naming Standards file are required."}), 400

    s2t_file = request.files['s2t_file']
    naming_standards_file = request.files['naming_standards_file']

    s2t_path = os.path.join(UPLOAD_FOLDER, s2t_file.filename)
    naming_standards_path = os.path.join(UPLOAD_FOLDER, naming_standards_file.filename)
    output_path = os.path.join(OUTPUT_FOLDER, "Processed_S2T.xlsx")

    try:
        logging.info("Saving uploaded files...")
        s2t_file.save(s2t_path)
        naming_standards_file.save(naming_standards_path)

        processed_file = process_files(s2t_path, naming_standards_path, output_path)
        logging.info("Processing completed successfully.")
        download_url = f"/download?file={os.path.basename(processed_file)}"
        return jsonify({"message": "Processing complete", "download_url": download_url}), 200

    except Exception as e:
        logging.error("Processing failed: %s", str(e))
        return jsonify({"error": str(e)}), 500

@app.route('/download', methods=['GET'])
def download_file():
    """Handles file download requests."""
    file_name = request.args.get('file')
    file_path = os.path.join(OUTPUT_FOLDER, file_name)  # Look for the file in the OUTPUT_FOLDER

    if not file_name or not os.path.exists(file_path):
        logging.error("File not found: %s", file_path)
        return jsonify({"error": "File not found."}), 404

    return send_file(file_path, as_attachment=True)

if __name__ == '__main__':
    logging.info("Starting Flask server...")
    app.run(debug=True)